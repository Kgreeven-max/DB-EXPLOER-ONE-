#!/usr/bin/env python3
"""
Populate MUIDs for Invalid Address Account List
Matches account numbers to MUIDs from fraudmonitor table
"""

import openpyxl
from db_connection import get_connection
from collections import defaultdict

# File paths
INPUT_FILE = r'C:\Users\kgreeven\Downloads\Invalid Addresses- Account List- MUIDS NEEDED 1 (1).xlsx'
OUTPUT_FILE = r'C:\Users\kgreeven\Downloads\Invalid Addresses- Account List- MUIDS POPULATED.xlsx'

def normalize_account_number(acc):
    """Normalize account number to 10-digit zero-padded format"""
    acc_str = str(acc).strip()
    # Handle non-numeric accounts (like "WC 12") - return as-is
    if not acc_str.replace(' ', '').isdigit():
        return acc_str
    # Zero-pad numeric accounts to 10 digits
    return acc_str.zfill(10)

def get_muid_mappings(account_numbers):
    """Query database to get all MUIDs for each account number"""
    if not account_numbers:
        return {}

    conn = get_connection()
    cursor = conn.cursor()

    # Normalize account numbers to match database format (10-digit zero-padded)
    normalized = [normalize_account_number(acc) for acc in account_numbers]

    # Build placeholders for IN clause
    placeholders = ', '.join(['%s'] * len(normalized))

    query = f"""
        SELECT DISTINCT masterMembership, muid
        FROM fraudmonitor
        WHERE masterMembership IN ({placeholders})
        AND muid IS NOT NULL
        AND muid != ''
        ORDER BY masterMembership, muid
    """

    cursor.execute(query, normalized)
    results = cursor.fetchall()

    cursor.close()
    conn.close()

    # Group MUIDs by normalized account number
    muid_map = defaultdict(list)
    for master_membership, muid in results:
        if muid not in muid_map[master_membership]:
            muid_map[master_membership].append(muid)

    # Create mapping from original account numbers to MUIDs
    original_to_muid = {}
    for orig, norm in zip(account_numbers, normalized):
        if norm in muid_map:
            original_to_muid[orig] = muid_map[norm]

    return original_to_muid

def main():
    print(f"Loading Excel file: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE)
    sheet = wb.active

    # Extract account numbers from Column A (skip header)
    account_numbers = []
    for row in range(2, sheet.max_row + 1):
        val = sheet.cell(row=row, column=1).value
        if val:
            account_numbers.append(str(val).strip())

    print(f"Found {len(account_numbers)} account numbers")

    # Query database for MUID mappings
    print("Querying database for MUID mappings...")
    muid_map = get_muid_mappings(account_numbers)

    # Count matches
    matched = sum(1 for acc in account_numbers if acc in muid_map)
    print(f"Found MUIDs for {matched} of {len(account_numbers)} accounts")

    # Populate columns B-E with MUIDs
    print("Populating Excel with MUIDs...")
    for row in range(2, sheet.max_row + 1):
        account = sheet.cell(row=row, column=1).value
        if account:
            account_str = str(account).strip()
            muids = muid_map.get(account_str, [])

            # Fill columns B-E (up to 4 MUIDs)
            for col_idx, muid in enumerate(muids[:4], start=2):
                sheet.cell(row=row, column=col_idx).value = muid

    # Save to new file
    print(f"Saving to: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print("Done!")

    # Print summary
    no_match = [acc for acc in account_numbers if acc not in muid_map]
    if no_match:
        print(f"\nAccounts with no MUID found ({len(no_match)}):")
        for acc in no_match[:20]:  # Show first 20
            print(f"  - {acc}")
        if len(no_match) > 20:
            print(f"  ... and {len(no_match) - 20} more")

if __name__ == "__main__":
    main()
